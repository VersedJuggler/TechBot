# tg_bot.py
import os
import asyncio
import tempfile
import json
import html
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from pathlib import Path
from dotenv import load_dotenv
from telegram.constants import ParseMode
load_dotenv()

import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    CallbackQueryHandler,
    filters,
)
import shutil

# ---------------------------------------------------------------------------
# Замените значение переменной на ваш токен или установите переменную
# окружения TG_BOT_TOKEN, чтобы токен подтянулся автоматически.
# ---------------------------------------------------------------------------
TOKEN: str | None = os.getenv("TG_BOT_TOKEN")
# Файл для хранения списка администраторов
ADMINS_FILE = "admins.json"

def _load_admins() -> set[int]:
    if os.path.exists(ADMINS_FILE):
        try:
            with open(ADMINS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return set(map(int, data.get("admins", [])))
        except Exception:
            pass
    # Если файла нет, возвращаем дефолтный набор (старые id)
    return {6413686861, 728567535, 510202114, 7548453140}

def _save_admins(admins: set[int]) -> None:
    try:
        with open(ADMINS_FILE, "w", encoding="utf-8") as f:
            json.dump({"admins": list(admins)}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def is_admin(user_id: int) -> bool:
    return user_id in _load_admins()


# Основные файлы для хранения
CATALOG_FILE = "catalog_data.json"
LATEST_EXCEL_FILE = "latest_catalog.xlsx"
MOVED_OVERRIDES_FILE = "moved_overrides.json"
def _load_moved_overrides() -> dict:
    if os.path.exists(MOVED_OVERRIDES_FILE):
        try:
            with open(MOVED_OVERRIDES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_moved_overrides(overrides: dict) -> None:
    try:
        with open(MOVED_OVERRIDES_FILE, "w", encoding="utf-8") as f:
            json.dump(overrides, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

MANUAL_CATEGORIES_FILE = "manual_categories.json"
def _load_manual_categories() -> dict:
    if os.path.exists(MANUAL_CATEGORIES_FILE):
        try:
            with open(MANUAL_CATEGORIES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_manual_categories(manual_cats: dict) -> None:
    try:
        with open(MANUAL_CATEGORIES_FILE, "w", encoding="utf-8") as f:
            json.dump(manual_cats, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    
    

# Названия кнопок главного меню
BTN_CHOOSE_CATEGORY = "🗂️ Выбор категории"
# Кнопка поиска по каталогу
BTN_CONTACT_MANAGER = "💬 Заказать товар у менеджера"
BTN_SUBSCRIBE = "✅ Подписаться"
BTN_GET_EXCEL = "💾 Получить Excel-файл"
BTN_SEARCH_CATALOG = "🔍 Поиск по каталогу"

# Добавим константу команды помощи
CMD_HELP = "help"

# Ссылки для связи с менеджером
MANAGER_TELEGRAM_LINK = "https://t.me/tanya_chilikova"
# Замените номер на актуальный формат для WhatsApp chat link
MANAGER_WHATSAPP_LINK = "https://wa.me/79278783209"


# Кнопка для админ-панели
BTN_ADMIN_PANEL = "🛠️ Админ-панель"

def get_main_menu_markup(is_admin=False):
    rows = [
        [BTN_CHOOSE_CATEGORY, BTN_CONTACT_MANAGER],
        [BTN_SUBSCRIBE, BTN_GET_EXCEL],
        [BTN_SEARCH_CATALOG],
    ]
    if is_admin:
        rows.append([BTN_ADMIN_PANEL])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)

MAIN_MENU_MARKUP = get_main_menu_markup(False)

# Порядок отображения категорий в основном меню
PREFERRED_CATEGORY_ORDER: list[str] = [
    "Телефоны",
    "Планшеты",
    "Ноутбуки",
]

def make_admin_panel_markup() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton("📥 Добавить каталог (.xlsx)", callback_data="adminpanel_add_catalog")],
        [InlineKeyboardButton("🔀 Изменить категорию товаров", callback_data="adminpanel_change_category")],
        [InlineKeyboardButton("📝 Ручные (manual)", callback_data="adminpanel_manual_root")],
        [InlineKeyboardButton("👤 Управление администраторами", callback_data="adminpanel_edit_admins")],
    ]
    return InlineKeyboardMarkup(buttons)

async def show_admin_panel(update_or_query, context):
    """
    Отправляет админ-панель как новое сообщение.
    update_or_query может быть и Update, и CallbackQuery.
    """
    chat_id = (
        update_or_query.effective_chat.id
        if hasattr(update_or_query, "effective_chat")
        else update_or_query.message.chat.id
    )
    await context.bot.send_message(
        chat_id=chat_id,
        text="🛠️ <b>Админ-панель</b>:",
        reply_markup=make_admin_panel_markup(),
        parse_mode="HTML"
    )


def _sort_categories(cat_names: list[str]) -> list[str]:
    """Возвращает список категорий в желаемом порядке отображения.

    1. Категории из PREFERRED_CATEGORY_ORDER – в указанной последовательности.
    2. Остальные (кроме "Другое") – по алфавиту.
    3. "Другое" – последней, если присутствует.
    """
    order_map = {name: idx for idx, name in enumerate(PREFERRED_CATEGORY_ORDER)}

    preferred = [c for c in PREFERRED_CATEGORY_ORDER if c in cat_names]
    other = sorted([c for c in cat_names if c not in order_map and c != "Другое"])
    tail = ["Другое"] if "Другое" in cat_names else []
    return preferred + other + tail



def _load_catalog_from_disk() -> dict | None:
    """Пытаемся загрузить каталог из файла JSON."""
    if os.path.exists(CATALOG_FILE):
        try:
            with open(CATALOG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return None

def get_full_catalog(context) -> dict:
    """Объединяет основной каталог, перенесённые товары и manual_categories для вывода и поиска."""
    catalog = context.application.bot_data.get("catalog") or {}
    moved = context.application.bot_data.get("moved_overrides") or {}
    manual = context.application.bot_data.get("manual_categories") or {}

    import copy
    # Глубокое копирование, чтобы не портить оригиналы
    full = copy.deepcopy(catalog)

    # Сначала добавляем перенесённые товары (moved_overrides)
    for cat, brands in moved.items():
        for brand, items in brands.items():
            full.setdefault(cat, {}).setdefault(brand, []).extend(copy.deepcopy(items))

    # Затем вручную добавленные товары (manual_categories)
    for cat, brands in manual.items():
        for brand, items in brands.items():
            full.setdefault(cat, {}).setdefault(brand, []).extend(copy.deepcopy(items))

    return full


def _save_catalog_to_disk(catalog: dict) -> None:
    """Сохраняем каталог в файл JSON."""
    try:
        with open(CATALOG_FILE, "w", encoding="utf-8") as f:
            json.dump(catalog, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        # При /start отменяем все промежуточные шаги ручного ввода
    for key in [
        "manualcat_step",
        "manualcat_category",
        "manualcat_brand",
        "manualcat_items",
        "manualcat_del_map",
        "manualprod_step",
        "manualprod_cat",
        "manualprod_brand",
        "manualprod_select_map",
        "manualprod_del_map",
        "change_step",
        "change_cat",
        "change_sub",
        "change_indices",
        "new_cat",
        "manualprice_step",
        "manualprice_cat",
        "manualprice_brand",
        "manualprice_indices",
        "manualprice_select_map",
    ]:
        context.user_data.pop(key, None)
        
    """Обработчик команды /start: приветствие и вывод каталога, если он загружен."""
    # Приветственное сообщение
    greet_text = (
        "Здравствуйте! Приветствуем вас в нашем каталоге. "
        "Вот что мы можем вам предложить"
    )

    user_id = update.effective_user.id if update.effective_user else None
    is_admin_user = user_id and is_admin(user_id)
    await update.message.reply_text(greet_text, reply_markup=get_main_menu_markup(is_admin_user))

    # Показать каталог, если он уже был загружен администратором
    # Используем объединённый каталог
    full_catalog = get_full_catalog(context)
    if full_catalog:
        buttons = []
        for cat_name in _sort_categories(list(full_catalog.keys())):
            subdict = full_catalog[cat_name]
            count = sum(len(items) for items in subdict.values())
            buttons.append([InlineKeyboardButton(text=f"{cat_name} ({count})", callback_data=f"cat|{cat_name}")])
        markup = InlineKeyboardMarkup(buttons)
        await update.message.reply_text("Выберите категорию:", reply_markup=markup)
    else:
        await update.message.reply_text("Каталог пока не загружен. Пожалуйста, попробуйте позже.")


# -------------------------------------------------------------------
# Правила классификации категорий и брендов (обновлено)
# -------------------------------------------------------------------


# Каждый элемент: (Категория, [список ключевых слов в нижнем регистре])
# Порядок — чем выше, тем выше приоритет.
CATEGORY_KEYWORDS: list[tuple[str, list[str]]] = [
    # Воздухоочистители
    ("Воздухоочистители", [
        "очиститель воздуха", "воздухоочиститель", "purifier"
    ]),
    # Отдельные специфичные категории → приоритет выше
    ("Телефоны противоударные", [
        "blackview", "doogee", "oukitel", "unihertz", "rugged", "armor", "tank", "cyber", "mega"
    ]),
    ("Телефоны кнопочные", ["nokia", "f+", "button phone", "feature phone"]),
    ("Игровые консоли", [
        "playstation", "ps4", "ps5", "xbox", "switch", "steam deck", "steamdeck",
        "джойстик", "игровая консоль", "игровая приставка"
    ]),
    ("VR-гарнитуры", [
    "vr", "vr шлем", "vr-шлем", "vr headset", "virtual reality",
    "oculus", "quest", "meta quest", "vive", "htc vive", "pico",
    "valve index", "reverb", "hp reverb", "ps vr", "psvr", "psvr2", "ps vr2"
    ]),
    (
        "Экшен-камеры",
        [
            # GoPro
            "gopro", "go pro", "hero", "gopro hero", "gopro hero 10", "gopro hero 11", "gopro hero 12",
            "gopro hero 13", "gopro hero 14", "gopro hero 15", "gopro hero 16", "gopro hero 17", "gopro hero 18", "gopro hero 19", "gopro hero 20",
            # DJI Osmo Action
            "dji", "osmo action", "action 5", "action5", "osmo action 5", "osmoaction",
            # Insta360
            "insta", "insta360", "insta 360"
        ],
    ),
    # Новая категория: Фен-стайлер (фены, стайлеры для волос)
    (
        "Фен-стайлер",
        [
            "фен",
            "стайлер",
            "фен-стайлер",
            "hair dryer",
            "styler",
            "airwrap",
            "supersonic",
            "hd08",
            "hd-08",
            "hd16",
            "hd-16",
            "hs08",
            "hs-08",
            "ht01",
            "ht-01",
        ],
    ),
    ("Пылесосы", ["пылесос", "vacuum", "робот-пылесос", "dyson", "dreame", "submarine"]),
    ("Планшеты", ["ipad", " galaxy tab", "tab ", "redmi pad", "poco pad", "tablet", "pad "]),
    ("Ноутбуки", ["ноутбук", "macbook", "magicbook", "matebook", "redmi book", "aspire", "ideapad", "ultrabook", "chromebook"]),
    ("Колонки", ["колонка", "speaker", "jbl", "marshall", "sber", "яндекс", "boombox", "partybox", "stanmore", "woburn", "макс"]),
    ("Наушники", ["наушник", "наушники", "airpods", "buds", "earphones", "earbuds", "sony wh-", "jbl tune", "marshall minor", "marshall major", "гарнитура"]),
    ("Часы", ["часы", "watch", "smart band", "galaxy fit", "fitbit", "amazfit", "gtr", "gt3"]),
    ("Телефоны", [
        "iphone", "samsung", "x.mi", "x.poco", "x.redmi", "honor", "google pixel", "zte", "realme",
        "oneplus", "asus zenfone", "смартфон", "smartphone", "galaxy"
    ]),
    ("Аксессуары", [
        "сзу", "сетевое зарядное устройство", "кабель", "переходник", "pencil", "keyboard", "mouse",
        "adapter", "magsafe", "беспроводная зарядка", "powerbank", "power bank", "чехол", "case", "cover"
    ]),
        ("Камеры видеонаблюдения", [
        "видеонаблюдени", "ip-камера", "ip камера", "cctv",
        "security camera", "ezviz", "hikvision", "dahua",
        "imou", "reolink", "wifi камера", "wi-fi камера", "tapo",
        "домашняя камера", "камера наблюдения"
    ]),
    ("Грили", [
        "гриль", "грили", "грильница", "электрогриль",
        "газовый гриль", "угольный гриль"
    ]),
    ("Квадрокоптеры", [
        "квадрокоптер", "квадрокоптеры", "коптер", "дрон",
        "drone", "quadcopter", "fpv", "mavic", "phantom", "air 2s", "mini 3", "mini 4"
    ]),
    ("Электроинструменты", [
        "шуруповерт", "шуруповёрт", "дрель", "перфоратор", "болгарка",
        "углошлифовальная", "лобзик", "пила", "шлифмашина", "фрезер",
        "реноватор", "сабельная пила", "гайковерт", "гайковёрт", "штроборез"
    ]),
    ("Бритвы, триммеры", [
        "бритва", "электробритва", "триммер", "машинка для стрижки",
        "шейвер", "shaver", "groom"
    ]),
    ("Эпиляторы", [
        "эпилятор", "фотоэпилятор", "ipl", "лазерная эпиляция"
    ]),
    ("Зубные щетки", [
        "зубная щетка", "зубные щетки", "электрическая щетка",
        "oral-b", "oral b", "sonicare", "oclean", "soocas", "щетка зубная", "щётка"
    ]),
]

BRAND_KEYWORDS: dict[str, str] = {
    # Смартфоны и электроника
    "apple": "Apple",
    "iphone": "Apple",
    "samsung": "Samsung",
    "galaxy": "Samsung",
    "xiaomi": "Xiaomi",
    "redmi": "Xiaomi",
    "poco": "Xiaomi",
    "mi ": "Xiaomi",
    "honor": "HONOR",
    "huawei": "Huawei",
    "google": "Google",
    "pixel": "Google",
    "zte": "ZTE",
    "realme": "Realme",
    "oneplus": "OnePlus",
    "asus": "ASUS",
    "zenfone": "ASUS",
    "lenovo": "Lenovo",
    "acer": "Acer",
    "gigabyte": "Gigabyte",
    "machenike": "Machenike",
    # Наушники и звук
    "jbl": "JBL",
    "marshall": "Marshall",
    "sony": "SONY",
    "sber": "Sber",
    "яндекс": "Яндекс",
    # Пылесосы и техника
    "dyson": "Dyson",
    "dreame": "Dreame",
    # Телефоны кнопочные / противоударные
    "nokia": "Nokia",
    "f+": "F+",
    "digma linx": "Digma Linx",
    "blackview": "Blackview",
    "doogee": "DOOGEE",
    "hotwav": "Hotwav",
    "oukitel": "OUKITEL",
    "unihertz": "Unihertz",
    # Прочее
    "gopro": "GoPro",
    "garmin": "Garmin",
    "fitbit": "Fitbit",
    # Экшен-камеры
    "dji": "DJI",
    "osmo": "DJI",
    "insta": "Insta360",
    "insta360": "Insta360",
    # VR 
    "oculus": "Oculus",
    "quest": "Oculus",
    "meta": "Meta",
    "htc": "HTC",
    "vive": "HTC",
    "pico": "Pico",
    "valve": "Valve",
    "valve index": "Valve",
    "hp": "HP",
    "reverb": "HP",         # HP Reverb
    "playstation": "SONY",  # чтобы PS VR/PS VR2 получили бренд SONY
    
    # Камеры видеонаблюдения
    "hikvision": "Hikvision",
    "dahua": "Dahua",
    "ezviz": "EZVIZ",
    "imou": "IMOU",
    "reolink": "Reolink",
    "tapo": "TP-Link Tapo",
    "tp-link": "TP-Link",
    "tplink": "TP-Link",

    # Квадрокоптеры
    "autel": "Autel",
    "hubsan": "Hubsan",
    "syma": "Syma",
    "parrot": "Parrot",

    # Электроинструменты
    "bosch": "Bosch",
    "makita": "Makita",
    "dewalt": "DeWALT",
    "de walt": "DeWALT",
    "metabo": "Metabo",
    "ryobi": "Ryobi",

    # Грили
    "weber": "Weber",
    "tefal": "Tefal",
    "redmond": "REDMOND",
    "kitfort": "Kitfort",
    "polaris": "Polaris",
    "george foreman": "George Foreman",

    # Бритвы, триммеры
    "philips": "Philips",
    "braun": "Braun",
    "panasonic": "Panasonic",
    "remington": "Remington",

    # Эпиляторы
    "rowenta": "Rowenta",

    # Зубные щетки
    "oral-b": "Oral-B",
    "oral b": "Oral-B",
    "sonicare": "Philips",
    "oclean": "Oclean",
    "soocas": "SOOCAS",
}


# --- Новая команда: /add_catalog ---
async def add_catalog_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда /add_catalog — загрузить новый Excel-файл с каталогом (только админ)."""
    user_id = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None
    if not user_id or not is_admin(user_id):
        if update.message:
            await update.message.reply_text("Извините, команда доступна только администратору.")
        elif chat_id:
            await context.bot.send_message(chat_id=chat_id, text="Извините, команда доступна только администратору.")
        return
    context.user_data["awaiting_file"] = True
    if update.message:
        await update.message.reply_text("Отправьте Excel-файл (.xlsx) с обновлённой базой товаров.")
    elif chat_id:
        await context.bot.send_message(chat_id=chat_id, text="Отправьте Excel-файл (.xlsx) с обновлённой базой товаров.")



# --- Новая команда: /edit_category ---
async def edit_category_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда /edit_category — управление вручную добавленными категориями (только админ)."""
    user_id = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None
    if not user_id or not is_admin(user_id):
        if update.message:
            await update.message.reply_text("Извините, команда доступна только администратору.")
        elif chat_id:
            await context.bot.send_message(chat_id=chat_id, text="Извините, команда доступна только администратору.")
        return
    # Загружаем вручную добавленные категории
    manual_cats = context.application.bot_data.get("manual_categories")
    if manual_cats is None:
        manual_cats = _load_manual_categories()
        context.application.bot_data["manual_categories"] = manual_cats
    if not manual_cats:
        msg = "Вручную добавленных категорий нет."
    else:
        lines = []
        for cat, brands in manual_cats.items():
            for brand, items in brands.items():
                lines.append(f"<b>{cat}</b> / <i>{brand}</i>: {len(items)} позиций")
        msg = "Вручную добавленные категории:\n" + "\n".join(lines)
    buttons = [
        [InlineKeyboardButton("Добавить", callback_data="manualcat_add")],
        [InlineKeyboardButton("Удалить", callback_data="manualcat_remove")],
    ]
    markup = InlineKeyboardMarkup(buttons)
    if update.message:
        await update.message.reply_text(msg, reply_markup=markup, parse_mode="HTML")
    elif chat_id:
        await context.bot.send_message(chat_id=chat_id, text=msg, reply_markup=markup, parse_mode="HTML")
    return

# --- Новая команда: /edit_products ---
async def edit_products_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда /edit_products — управление товарами внутри вручную добавленных подкатегорий (только админ)."""
    user_id = update.effective_user.id if update.effective_user else None
    if not user_id or not is_admin(user_id):
        await update.message.reply_text("Извините, команда доступна только администратору.")
        return

    # Загружаем вручную добавленные категории/бренды
    manual_cats = context.application.bot_data.get("manual_categories")
    if manual_cats is None:
        manual_cats = _load_manual_categories()
        context.application.bot_data["manual_categories"] = manual_cats

    # Проверка наличия вручную добавленных подкатегорий
    items = []
    cb_map = {}
    idx = 0
    for cat, brands in manual_cats.items():
        for brand in brands:
            cb_key = f"manualprod_select|{idx}"
            cb_map[cb_key] = (cat, brand)
            items.append([InlineKeyboardButton(f"{cat} / {brand}", callback_data=cb_key)])
            idx += 1

    if not items:
        await update.message.reply_text("Нет вручную добавленных подкатегорий для управления товарами.")
        return

    context.user_data["manualprod_select_map"] = cb_map
    markup = InlineKeyboardMarkup(items)
    await update.message.reply_text("Выберите подкатегорию для редактирования товаров:", reply_markup=markup)

# --- Новая команда: /edit_admins ---
async def edit_admins_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда /edit_admins — управление списком администраторов (только админ)."""
    user_id = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None
    if not user_id or not is_admin(user_id):
        if update.message:
            await update.message.reply_text("Извините, команда доступна только администратору.")
        elif chat_id:
            await context.bot.send_message(chat_id=chat_id, text="Извините, команда доступна только администратору.")
        return
    # Показываем список админов и две кнопки: Добавить, Удалить
    admins = _load_admins()
    admin_lines = []
    for admin_id in admins:
        try:
            user = await context.bot.get_chat(admin_id)
            username = f"@{user.username}" if getattr(user, "username", None) else ""
        except Exception:
            username = ""
        admin_lines.append(f"{admin_id} {username}")
    msg = (
        "Текущие администраторы:\n"
        + "\n".join(admin_lines)
    )
    buttons = [
        [InlineKeyboardButton("Добавить", callback_data="admin_add")],
        [InlineKeyboardButton("Удалить", callback_data="admin_remove")],
    ]
    markup = InlineKeyboardMarkup(buttons)
    if update.message:
        await update.message.reply_text(msg, reply_markup=markup)
    elif chat_id:
        await context.bot.send_message(chat_id=chat_id, text=msg, reply_markup=markup)
    return


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда /help — выводит информацию о связи с менеджером."""
    help_text = (
        "📦 <b>Как оформить заказ:</b>\n\n"
        "Нажмите «💬 Заказать товар у менеджера»\n\n"
        "<b>В сообщении укажите</b> точную модель товара, который вас интересует "
        "(например: <i>MacBook Pro 16 M4, 24/512, Black</i>)\n\n"
        "Мы подтвердим наличие и <b>зарезервируем</b> товар за вами\n\n"
        "🚚 <b>Доставка по Москве:</b>\n\n"
        "В пределах МКАД — от <b>1 000 ₽</b>\n"
        "За МКАД (до 30 км) — по договорённости\n\n"
        "🛍 <b>Самовывоз — бесплатно:</b>\n\n"
        "Заказы, оформленные до <b>13:00</b>, можно получить в тот же день\n"
        "После <b>13:00</b> — на следующий день\n\n"
        "🕒 <b>Выдача заказов:</b>\n"
        "⏰ Ежедневно с <b>15:00</b> до <b>16:00</b>\n"
        "📍 Адрес: <b>ТЦ Рубин, Багратионовский проезд, 7к2</b>\n"
        "(5 минут пешком от метро <i>Багратионовская</i>)"
    )
    back_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton(text="← Назад", callback_data="back|root")]
    ])
    await update.message.reply_text(
        help_text,
        reply_markup=back_markup,
        parse_mode=ParseMode.HTML
    )

async def about_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Команда /about — краткая информация о магазине + кнопка Назад."""
    text = (
        "<b>V&amp;P Tech</b> — оригинальная техника и электроника по низким ценам.\n"
        "📦 Всё в наличии, с гарантией.\n"
        "🚚 В Москве — доставка или самовывоз в день заказа.\n"
        "📬 По России — отправляем СДЭК, Яндекс, Почтой.\n"
        "✅ Работаем давно.\n"
        "💬 Нужна помощь? Менеджер всегда на связи!"
    )
    back_markup = InlineKeyboardMarkup([
        [InlineKeyboardButton(text="← Назад", callback_data="back|root")]
    ])
    await update.message.reply_text(text, reply_markup=back_markup, parse_mode=ParseMode.HTML)


import re

def extract_category(description: str) -> tuple[str, str]:
    """
    Категоризация товара по описанию с учетом приоритетов, перекрестных признаков и гибких правил.
    """
    desc = description or ""
    desc_low = desc.lower()
    category = "Другое"
    subcategory = "Общее"

    # 1. Воздухоочистители (бренды: Xiaomi, Dyson, Philips, Sharp, Boneco, Levoit)
    if any(x in desc_low for x in ["очиститель воздуха", "воздухоочиститель", "purifier"]):
        for kw, brand in [
            ("xiaomi", "Xiaomi"),
            ("dyson", "Dyson"),
            ("philips", "Philips"),
            ("sharp", "Sharp"),
            ("boneco", "Boneco"),
            ("levoit", "Levoit")
        ]:
            if kw in desc_low:
                return "Воздухоочистители", brand
        return "Воздухоочистители", "Общее"

    # 2. Игровые консоли: SteamDeck как отдельный бренд
    if "steam deck" in desc_low or "steamdeck" in desc_low:
        return "Игровые консоли", "SteamDeck"

    # 3. Исключить Mi TV Box из телефонов/Xiaomi
    if ("mi tv box" in desc_low or "xiaomi tv box" in desc_low) and ("телефон" in desc_low or "xiaomi" in desc_low):
        return "Другое", "Общее"


    # --- 1. Наушники (приоритет: явное слово, AirPods, EarPods, Buds, Earphones, Earbuds, гарнитура, даже если есть type-c, usb-c и т.д.) ---
    headphones_pattern = r"\b(наушник|наушники|airpods|air pods|air pod|earpods|ear pods|ear pod|earphones|earphone|earbuds|earbud|buds|гарнитура)\b"
    if re.search(headphones_pattern, desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Наушники", brand
        return "Наушники", "Общее"


    # --- 2. Планшеты (Pad, Tab, Tablet, кроме Notepad) ---
    # Гибкий паттерн: tab, tablet, pad, galaxy tab, redmi pad, poco pad, ipad, и т.д.
    tablet_pattern = r"(ipad|\btab\b|tablet|pad(?![a-z]))"
    if (re.search(tablet_pattern, desc_low) or re.search(r"pad[\s\d]", desc_low)) and not re.search(r"notepad", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Планшеты", brand
        return "Планшеты", "Общее"

    # --- 3. Явные аксессуары (расширено) ---
    accessories_kw = [
        "аксессуар", "чехол", "стекло", "кабель", "шнур", "переходник", "adapter", "зарядка", "powerbank", "power bank", "magsafe", "pencil", "cover", "case", "screen protector", "беспроводная зарядка", "сетевое зарядное устройство", "сзу", "блок", "адаптер", "блок питания", "usb", "type-c", "lightning", "micro-usb", "магнитный кабель", "стекло защитное", "защитное стекло", "док-станция", "док станция", "док", "hub", "разветвитель", "splitter", "держатель", "mount", "подставка", "ремешок", "strap", "ремень", "пленка", "film", "наклейка", "наклейки", "stylus", "стилус"
    ]
    if any(re.search(rf"(?<![а-яa-z0-9]){re.escape(kw)}(?![а-яa-z0-9])", desc_low) for kw in accessories_kw):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Аксессуары", brand
        return "Аксессуары", "Общее"

    # --- 3. Колонки (исключая наушники) ---
    if re.search(r"\b(колонка|speaker|boombox|partybox|stanmore|woburn)\b", desc_low) and not re.search(r"наушник|наушники|buds|earbuds|гарнитура", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Колонки", brand
        return "Колонки", "Общее"


    # --- 4. Фен-стайлеры (Dyson, Supersonic, Airwrap и др.) ---
    if re.search(r"фен|стайлер|hair dryer|styler|airwrap|supersonic|hd08|hd-08|hd16|hd-16|hs08|hs-08|ht01|ht-01", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Фен-стайлер", brand
        return "Фен-стайлер", "Общее"

    # --- 5. Пылесосы (все бренды, любые слова) ---
    # Паттерн: пылесос, vacuum, cleaner, робот-пылесос, robot vacuum, robot cleaner, робот vacuum, робот cleaner, dreame, dyson, submarine
    if re.search(r"пылесос|vacuum|cleaner|робот-пылесос|robot vacuum|robot cleaner|робот vacuum|робот cleaner|dreame|dyson|submarine", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Пылесосы", brand
        return "Пылесосы", "Общее"

    # --- 5. Часы и браслеты (Garmin, Band, Instinct и др.) ---
    if re.search(r"\b(часы|watch|band|fitbit|amazfit|gtr|gt3|instinct|forerunner|fenix|coros|garmin|band)\b", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Часы", brand
        return "Часы", "Общее"

    # --- 6. Планшеты (Pad, Tab, Tablet, кроме Notepad) ---
    if (re.search(r"\bipad\b|\btab\b|\btablet\b|\bpad\b", desc_low) or re.search(r"pad[\s\d]", desc_low)) and not re.search(r"notepad", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Планшеты", brand
        return "Планшеты", "Общее"

    # --- 7. Ноутбуки (Apple, Matebook, CPU, дюймы, модели, book, клавиатура) ---
    # Явные признаки ноутбука: 'book' + дюймы, или 'клавиатура' (RU клавиатура и др.)
    if (re.search(r"book", desc_low) and re.search(r"\d{2}\"", desc)) or re.search(r"клавиатура", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Ноутбуки", brand
        return "Ноутбуки", "Общее"
    # Apple MacBook: Air/Pro + 13"/14"/15"/16"/M1/M2/M3/M4
    if (re.search(r"macbook|air|pro", desc_low) and (re.search(r"\d{2}\"", desc) or re.search(r"\bm[1-4]\b", desc_low))) or re.search(r"macbook", desc_low):
        return "Ноутбуки", "Apple"
    # Matebook, ноутбуки других брендов
    if re.search(r"matebook|notebook|ultrabook|chromebook|magicbook|aspire|ideapad|thinkpad|vivobook|zenbook|legion|gigabyte|machenike|lenovo|acer|asus|hp|dell|msi|huawei", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Ноутбуки", brand
        return "Ноутбуки", "Общее"
    # Intel/AMD CPU + 13"/14"/15"/16"
    if re.search(r"(intel|amd|ryzen|core i[3579]|pentium|celeron)", desc_low) and re.search(r"\d{2}\"", desc):
        return "Ноутбуки", "Общее"

    # --- 7. Телефоны (Mate X, бренды, явные признаки) ---
    # Huawei Mate X6 — телефон, Matebook — ноутбук
    if re.search(r"matebook", desc_low):
        return "Ноутбуки", "Huawei"
    if re.search(r"mate", desc_low) and not re.search(r"matebook", desc_low):
        return "Телефоны", "Huawei"
    # Смартфоны по брендам и ключевым словам
    # Исключить Mi TV Box из телефонов/Xiaomi (ещё раз для надёжности)
    if ("mi tv box" in desc_low or "xiaomi tv box" in desc_low):
        return "Другое", "Общее"
    phone_kw = ["iphone", "смартфон", "smartphone", "galaxy", "pixel", "zenfone", "oneplus", "realme", "zte", "redmi", "poco", "xiaomi", "samsung", "huawei", "honor"]
    if any(re.search(rf"(?<![а-яa-z0-9]){re.escape(kw)}(?![а-яa-z0-9])", desc_low) for kw in phone_kw):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                # Исключить Mi TV Box из телефонов/Xiaomi
                if brand == "Xiaomi" and ("mi tv box" in desc_low or "xiaomi tv box" in desc_low):
                    return "Другое", "Общее"
                return "Телефоны", brand
        return "Телефоны", "Общее"


    # --- 8. Кнопочные телефоны ---
    if re.search(r"button phone|feature phone|nokia|f\+|digma linx", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Телефоны кнопочные", brand
        return "Телефоны кнопочные", "Общее"

    # --- 9. Противоударные телефоны ---
    if re.search(r"противоударный|rugged|armor|tank|cyber|mega|blackview|doogee|hotwav|oukitel|unihertz", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Телефоны противоударные", brand
        return "Телефоны противоударные", "Общее"

    # --- НОВОЕ: VR-гарнитуры ---
    if re.search(r"(?:\bvr\b|vr-?шлем|vr\s?headset|virtual\s+reality|meta\s?quest|oculus|quest(?:\s?(?:2|3|pro))?|htc\s?vive|(?:^|\b)vive\b|pico|valve\s?index|hp\s?reverb|reverb\s?g2|ps\s?vr2?|psvr2?)", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "VR-гарнитуры", brand
        return "VR-гарнитуры", "Общее"
    
    # --- 9. Игровые консоли (без VR) ---
    if re.search(r"playstation|ps4|ps5|xbox|switch|steam deck|steamdeck|джойстик|игровая консоль|игровая приставка", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Игровые консоли", brand
        return "Игровые консоли", "Общее"
    
        # --- НОВОЕ: Камеры видеонаблюдения ---
    if re.search(r"(видеонаблюдени|ip[-\s]?камера|cctv|security camera|wi-?fi\s?камера|домашняя камера|ezviz|hikvision|dahua|imou|reolink|tapo)", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Камеры видеонаблюдения", brand
        return "Камеры видеонаблюдения", "Общее"

    # --- НОВОЕ: Квадрокоптеры ---
    if re.search(r"\b(квадро?коптеры?|коптер|дрон|drone|quadcopter|fpv)\b", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Квадрокоптеры", brand
        return "Квадрокоптеры", "Общее"

    # --- НОВОЕ: Грили ---
    if re.search(r"\b(гриль|грили|грильница|электрогриль|газовый гриль|угольный гриль)\b", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Грили", brand
        return "Грили", "Общее"

    # --- НОВОЕ: Электроинструменты ---
    if re.search(r"\b(шуруповёрт|шуруповерт|дрель|перфоратор|болгарка|углошлифовальная|лобзик|пила|шлифмашин|фрезер|реноватор|сабельная пила|гайковёрт|гайковерт|штроборез)\b", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Электроинструменты", brand
        return "Электроинструменты", "Общее"

    # --- НОВОЕ: Бритвы, триммеры ---
    if re.search(r"\b(бритва|электробритва|триммер|машинка для стрижки|шейвер|shaver|groom)\b", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Бритвы, триммеры", brand
        return "Бритвы, триммеры", "Общее"

    # --- НОВОЕ: Эпиляторы ---
    if re.search(r"\b(эпилятор|фотоэпилятор|ipl|лазерн\w*\sэпиляц\w*)\b", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Эпиляторы", brand
        return "Эпиляторы", "Общее"

    # --- НОВОЕ: Зубные щетки ---
    if re.search(r"(зубн\w*\sщ(е|ё)тка|электрическ\w*\sщ(е|ё)тка|oral-?b|sonicare|oclean|soocas)", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Зубные щетки", brand
        return "Зубные щетки", "Общее"

    # --- 10. Экшен-камеры ---
    if re.search(r"gopro|osmo action|insta360|insta 360|dji|hero", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                 return "Экшен-камеры", brand
        return "Экшен-камеры", "Общее"

    # --- 11. Фен-стайлеры ---
    if re.search(r"фен|стайлер|hair dryer|styler|airwrap|supersonic|hd08|hd-08|hd16|hd-16|hs08|hs-08|ht01|ht-01", desc_low):
        return "Фен-стайлер", "Общее"

    # --- 12. Пылесосы ---
    if re.search(r"пылесос|vacuum|робот-пылесос|dyson|dreame|submarine", desc_low):
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                return "Пылесосы", brand
        return "Пылесосы", "Общее"  

    # --- 13. Категория по ключевым словам (fallback) ---
    for cat, keywords in CATEGORY_KEYWORDS:
        if any(kw in desc_low for kw in keywords):
            category = cat
            break

    # --- 14. Бренд по ключевым словам (fallback) ---
    first_word = desc.split()[0].strip(',.;:"()').lower() if desc else ""
    if first_word and first_word in BRAND_KEYWORDS:
        subcategory = BRAND_KEYWORDS[first_word]
    else:
        for kw, brand in BRAND_KEYWORDS.items():
            if kw in desc_low:
                subcategory = brand
                break

    # --- 15. Особое правило: для категории Go Pro всегда бренд GoPro ---
    if category == "Go Pro":
        subcategory = "GoPro"

    return category, subcategory


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None: 
    """При получении документа проверяем, что это .xlsx, скачиваем и обрабатываем."""
    user_id = update.effective_user.id if update.effective_user else None
    awaiting_file = context.user_data.get("awaiting_file")
    if not user_id or not is_admin(user_id) or not awaiting_file:
        await update.message.reply_text(
            "Извините, сейчас бот не ожидает файл или у вас нет прав загрузки."
        )
        return

    # Сбрасываем флаг ожидания файла
    context.user_data["awaiting_file"] = False

    document = update.message.document
    if not document:
        return

    if not document.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text(
            "Пожалуйста, отправьте файл в формате .xlsx. Другие форматы не поддерживаются."
        )
        return

    # Сохраняем файл во временную директорию
    tmp_dir = Path(tempfile.mkdtemp())
    src_path = tmp_dir / document.file_name
    file_obj = await document.get_file()
    await file_obj.download_to_drive(str(src_path))

    try:
        # Читаем Excel
        df = pd.read_excel(src_path)
    except Exception as exc:
        await update.message.reply_text(
            "Не удалось прочитать файл как Excel: " f"{exc}"
        )
        return

    # Сохраняем копию файла, чтобы пользователи могли скачивать актуальную версию
    try:
        shutil.copy(src_path, LATEST_EXCEL_FILE)
    except Exception:
        pass

    # Строим каталог по описанию
    catalog: dict[str, dict[str, list[dict[str, str]]]] = {}
    for _, row in df.iterrows():
        desc = str(row.get("description") or row.get("desription") or "")
        price = row.get("price") or row.get("Цена") or row.get("Price") or ""
        cat, sub = extract_category(desc)
        catalog.setdefault(cat, {}).setdefault(sub, []).append({"desc": desc, "price": price})

    # === СИНХРОНИЗАЦИЯ ПЕРЕНЕСЁННЫХ (moved_overrides) С EXCEL И УБОРКА ДУБЛЕЙ ===
    def _norm_desc(s: str) -> str:
        import re as _re
        return _re.sub(r"\s+", " ", str(s or "").strip().lower())

    # 1) Карта "нормализованное описание -> цена" из Excel
    excel_price_by_desc: dict[str, str] = {}
    for _, row in df.iterrows():
        d = str(row.get("description") or row.get("desription") or "")
        p = row.get("price") or row.get("Цена") or row.get("Price") or ""
        excel_price_by_desc[_norm_desc(d)] = p

    # 2) Обновляем цены в moved_overrides и удаляем те, которых больше нет в Excel
    overrides = context.application.bot_data.get("moved_overrides")
    if overrides is None:
        overrides = _load_moved_overrides()

    changed = False
    to_del_cats = []
    for cat, brands in list(overrides.items()):
        to_del_brands = []
        for brand, items in list(brands.items()):
            new_items = []
            for it in items:
                key = _norm_desc(it.get("desc", ""))
                if key in excel_price_by_desc:
                    new_price = excel_price_by_desc[key]
                    if str(it.get("price", "")) != str(new_price):
                        it["price"] = new_price
                        changed = True
                    new_items.append(it)
                else:
                    # Позиции больше нет в Excel -> удаляем из перенесённых
                    changed = True
            if new_items:
                overrides[cat][brand] = new_items
            else:
                to_del_brands.append(brand)
        for b in to_del_brands:
            del overrides[cat][b]
        if not overrides.get(cat):
            to_del_cats.append(cat)
    for c in to_del_cats:
        del overrides[c]

    if changed:
        _save_moved_overrides(overrides)
        context.application.bot_data["moved_overrides"] = overrides

    # 3) Убираем из авто-каталога все позиции, что уже есть в moved_overrides ИЛИ manual_categories
    manual = context.application.bot_data.get("manual_categories") or _load_manual_categories()

    occupied_descs = {
        _norm_desc(mi.get("desc", ""))
        for source in (overrides, manual)
        for brands in source.values()
        for sublist in brands.values()
        for mi in sublist
    }

    for cat_key in list(catalog.keys()):
        for sub_key in list(catalog[cat_key].keys()):
            filtered = [
                item for item in catalog[cat_key][sub_key]
                if _norm_desc(item.get("desc", "")) not in occupied_descs
            ]
            if filtered:
                catalog[cat_key][sub_key] = filtered
            else:
                del catalog[cat_key][sub_key]
        if not catalog.get(cat_key):
            del catalog[cat_key]
    # === КОНЕЦ СИНХРОНИЗАЦИИ ===

    if not catalog:
        await update.message.reply_text("Не удалось сформировать категории по описанию.")
        return

    # Сохраняем каталог в bot_data (общий для всех пользователей)
    context.application.bot_data["catalog"] = catalog
    # А также на диск, чтобы каталог сохранялся между перезапусками бота
    _save_catalog_to_disk(catalog)

    # После успешной загрузки каталога выводим сообщение с инструкцией
    await update.message.reply_text("Каталог успешно добавлен, нажмите /start, чтобы ознакомиться с категориями")

    # Удаляем временный файл
    try:
        os.remove(src_path)
        os.rmdir(tmp_dir)
    except OSError:
        pass


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Обработка текстовых сообщений и нажатий на кнопки меню."""
    import re

    text = update.message.text
    user_id = update.effective_user.id if update.effective_user else None
    is_admin_user = user_id and is_admin(user_id)

    # --- Изменение цен: шаг 1 — ввод номеров строк ---
    if context.user_data.get("manualprice_step") == "awaiting_indices":
        raw = (text or "").strip()
        parts = re.split(r"[,\s]+", raw)
        idxs = set()
        try:
            for part in parts:
                if not part:
                    continue
                if "-" in part:
                    a, b = map(int, part.split("-", 1))
                    if a > b:
                        a, b = b, a
                    idxs.update(range(a, b + 1))
                else:
                    idxs.add(int(part))
        except Exception:
            await update.message.reply_text("Некорректный формат. Пример: 1-3,5")
            return
    
        indices = sorted({i - 1 for i in idxs if i > 0})
        if not indices:
            await update.message.reply_text("Не выбраны строки. Укажите номера, например: 1-3,5")
            return
    
        context.user_data["manualprice_indices"] = indices
        context.user_data["manualprice_step"] = "awaiting_price"
        await update.message.reply_text("Введите новую цену (одно значение будет применено ко всем выбранным товарам):")
        return
    
    # --- Изменение цен: шаг 2 — ввод новой цены и сохранение ---
    if context.user_data.get("manualprice_step") == "awaiting_price":
        new_price = (text or "").strip()
        if not new_price:
            await update.message.reply_text("Цена не может быть пустой. Введите новое значение.")
            return
    
        cat = context.user_data.pop("manualprice_cat", None)
        brand = context.user_data.pop("manualprice_brand", None)
        indices = context.user_data.pop("manualprice_indices", [])
        context.user_data.pop("manualprice_step", None)
    
        manual = context.application.bot_data.get("manual_categories", {}) or _load_manual_categories()
        items = manual.get(cat, {}).get(brand, [])
    
        updated = 0
        for i in indices:
            if 0 <= i < len(items):
                items[i]["price"] = new_price
                items[i]["price_locked"] = True
                updated += 1
    
        _save_manual_categories(manual)
        context.application.bot_data["manual_categories"] = manual
    
        await update.message.reply_text(f"✅ Обновлено цен: {updated} шт. в {cat} / {brand}.")
        # вернёмся в админ-панель (если у вас уже есть вспомогательная функция)
        try:
            await show_admin_panel(update, context)
        except NameError:
            pass
        return

    # --- 0.1. Пошаговое добавление вручную категории/бренда/товаров ---
    if context.user_data.get("manualcat_step"):
        step = context.user_data["manualcat_step"]
        user_id = update.effective_user.id if update.effective_user else None
        if not user_id or not is_admin(user_id):
            await update.message.reply_text("Нет прав для добавления.")
            context.user_data.pop("manualcat_step", None)
            return

        if step == 1:
            # Получили название категории
            cat = text.strip()
            if not cat:
                await update.message.reply_text("Название категории не может быть пустым. Введите ещё раз:")
                return
            context.user_data["manualcat_category"] = cat
            context.user_data["manualcat_step"] = 2
            await update.message.reply_text("Введите название бренда (подкатегории):")
            return

        elif step == 2:
            # Получили название бренда
            brand = text.strip()
            if not brand:
                await update.message.reply_text("Название бренда не может быть пустым. Введите ещё раз:")
                return
            context.user_data["manualcat_brand"] = brand
            context.user_data["manualcat_step"] = 3
            await update.message.reply_text(
                "Введите описание товара и цену.\nКаждая строка: Описание;Цена\n\n Для создания пустой категории введите '0'.\n\n"
            )
            context.user_data["manualcat_items"] = []
            return

        elif step == 3:
                        # ——— Если ввели "0" — создаём пустую категорию и выходим ———
            if text.strip() == "0":
                cat = context.user_data.pop("manualcat_category")
                brand = context.user_data.pop("manualcat_brand")
                context.user_data.pop("manualcat_step", None)

                # Загрузить или инициализировать manual_categories
                manual_cats = context.application.bot_data.get("manual_categories")
                if manual_cats is None:
                    manual_cats = _load_manual_categories()

                # Создать пустой список товаров в новой подкатегории
                manual_cats.setdefault(cat, {})[brand] = []
                context.application.bot_data["manual_categories"] = manual_cats
                _save_manual_categories(manual_cats)

                # Ответить администратору
                buttons = [
                    [InlineKeyboardButton("Добавить ещё", callback_data="manualcat_add")],
                    [InlineKeyboardButton("← Назад", callback_data="manualcat_remove")]
                ]
                markup = InlineKeyboardMarkup(buttons)
                await update.message.reply_text(
                    f"✅ Создана пустая категория: <b>{cat}</b> / <i>{brand}</i>.\n\n"
                    "Теперь вы можете добавить в неё товары или перенести что-то позже.",
                    reply_markup=markup,
                    parse_mode="HTML"
                )
                await show_admin_panel(update, context)
                return
            
            # Получаем товары (многострочно, до 'Готово')
            if text.strip().lower() == "готово":
                await update.message.reply_text(
                    "Пожалуйста, отправьте список товаров (каждая строка: Описание;Цена). "
                    "Если хотите отменить — используйте /start."
                )
                return

            # Ожидаем список товаров, каждая строка: Описание;Цена
            lines = [line for line in text.splitlines() if line.strip()]
            items = []
            for line in lines:
                parts = line.split(";")
                if len(parts) < 2:
                    continue  # пропускаем некорректные строки
                desc = parts[0].strip()
                price = parts[1].strip()
                if not desc or not price:
                    continue
                items.append({"desc": desc, "price": price, "price_locked": True, "origin": "manual"})

            if items:
                cat = context.user_data.pop("manualcat_category")
                brand = context.user_data.pop("manualcat_brand")
                context.user_data.pop("manualcat_step", None)

                # Сохраняем в manual_categories.json
                manual_cats = context.application.bot_data.get("manual_categories")
                if manual_cats is None:
                    manual_cats = _load_manual_categories()
                manual_cats.setdefault(cat, {}).setdefault(brand, []).extend(items)
                context.application.bot_data["manual_categories"] = manual_cats
                _save_manual_categories(manual_cats)

                # Показываем обновлённый список вручную добавленных категорий
                lines = []
                for c, brands in manual_cats.items():
                    for b, its in brands.items():
                        lines.append(f"<b>{c}</b> / <i>{b}</i>: {len(its)} позиций")
                msg = "Вручную добавленные категории:\n" + "\n".join(lines)
                buttons = [
                    [InlineKeyboardButton("Добавить", callback_data="manualcat_add")],
                    [InlineKeyboardButton("Удалить", callback_data="manualcat_remove")],
                ]
                markup = InlineKeyboardMarkup(buttons)
                await update.message.reply_text(
                    f"Добавлено в {cat} / {brand}: {len(items)} позиций.\n\n{msg}",
                    reply_markup=markup,
                    parse_mode="HTML"
                )
                await show_admin_panel(update, context)
            else:
                await update.message.reply_text(
                    "Не удалось добавить ни одного товара. Проверьте формат: Описание;Цена."
                )
            return

    # --- 0.2. Обработка шагов добавления товаров в существующую подкатегорию ---
    if context.user_data.get("manualprod_step") and context.user_data["manualprod_step"] == 1:
        # Сбор введённых строк при добавлении товаров
        text_in = update.message.text
        if text_in.strip().lower() == "готово":
            await update.message.reply_text(
                "Добавление отменено или завершено неверно. Начните заново."
            )
            context.user_data.pop("manualprod_step", None)
            return

        lines = [l for l in text_in.splitlines() if l.strip()]
        items = []
        for line in lines:
            parts = line.split(";")
            if len(parts) < 2:
                continue
            desc, price = parts[0].strip(), parts[1].strip()
            if desc and price:
                items.append({"desc": desc, "price": price, "price_locked": True, "origin": "manual"})

        if items:
            cat = context.user_data.pop("manualprod_cat")
            brand = context.user_data.pop("manualprod_brand")
            context.user_data.pop("manualprod_step", None)

            manual_cats = context.application.bot_data.get("manual_categories") or _load_manual_categories()
            manual_cats.setdefault(cat, {}).setdefault(brand, []).extend(items)
            _save_manual_categories(manual_cats)
            context.application.bot_data["manual_categories"] = manual_cats

            await update.message.reply_text(
                f"Добавлено в {cat} / {brand}: {len(items)} позиций."
            )
            await show_admin_panel(update, context)
        else:
            await update.message.reply_text(
                "Не удалось разобрать ни одну строку. Проверьте формат: Описание;Цена."
            )
        return

    # --- Шаг удаления товаров по вводимым номерам ---
    if context.user_data.get("manualprod_step") == "awaiting_manualprod_delete":
        raw = update.message.text.strip()
        parts = re.split(r"[,\s]+", raw)
        idxs = set()
        for part in parts:
            if "-" in part:
                a, b = map(int, part.split("-", 1))
                idxs.update(range(a, b + 1))
            else:
                idxs.add(int(part))

        # Переводим в 0-based и сортируем по убыванию, чтобы удалять корректно
        indices = sorted({i - 1 for i in idxs if i > 0}, reverse=True)

        # Достаём контекст
        cat = context.user_data.pop("manualprod_cat", None)
        brand = context.user_data.pop("manualprod_brand", None)
        context.user_data.pop("manualprod_step", None)

        manual = context.application.bot_data.get("manual_categories", {}) or _load_manual_categories()
        items = manual.get(cat, {}).get(brand, [])

        removed = []
        for i in indices:
            if 0 <= i < len(items):
                removed.append(items.pop(i))

        # Сохраняем изменения
        _save_manual_categories(manual)
        context.application.bot_data["manual_categories"] = manual

        if removed:
            lines = []
            for it in removed:
                d = html.escape(it.get("desc", ""))
                p = html.escape(str(it.get("price", "")))
                lines.append(f"— {d} ({p})")
            await update.message.reply_text(
                "<b>Удалено товаров:</b> {}\n\n{}".format(len(removed), "\n".join(lines)),
                parse_mode=ParseMode.HTML
            )
            await show_admin_panel(update, context)
        else:
            await update.message.reply_text("Ничего не удалено (неверные номера).")
        return

    # --- 0.3. Ожидание ввода user_id для добавления/удаления админа ---
    if context.user_data.get("awaiting_admin_action"):
        action = context.user_data.pop("awaiting_admin_action")
        user_id = update.effective_user.id if update.effective_user else None
        if not user_id or not is_admin(user_id):
            await update.message.reply_text("Нет прав для изменения админов.")
            return
        try:
            target_id = int(text.strip())
        except ValueError:
            await update.message.reply_text("user_id должен быть числом.")
            return

        admins = _load_admins()
        if action == "add":
            admins.add(target_id)
            _save_admins(admins)
            await update.message.reply_text(f"Пользователь {target_id} добавлен в администраторы.")
        elif action == "remove":
            if target_id in admins:
                admins.remove(target_id)
                _save_admins(admins)
                await update.message.reply_text(f"Пользователь {target_id} удалён из администраторов.")
            else:
                await update.message.reply_text("Такого пользователя нет в списке админов.")
        return

    # --- 1. Обработка режима поиска ---
    if context.user_data.pop("awaiting_search", False):
        # 1) Нормализуем запрос
        raw = (text or "").strip()
        if not raw:
            await update.message.reply_text("Пустой запрос. Попробуйте ещё раз.")
            return
        q = raw.lower()
        q = re.sub(r'([a-zа-яё])(\d)', r'\1 \2', q)
        q = re.sub(r'(\d)([a-zа-яё])', r'\1 \2', q)

        # 2.1) СПЕЦ-СЛУЧАЙ: «macbook» и его вариации → только Ноутбуки / Apple
        mac = q.replace(" ", "")
        if mac.startswith("macbook"):
            full_catalog = get_full_catalog(context)
            results = [
                ("Ноутбуки", "Apple", item)
                for item in full_catalog.get("Ноутбуки", {}).get("Apple", [])
            ]
            if not results:
                await update.message.reply_text("Ничего не найдено по вашему запросу.")
                return

            await update.message.reply_text(f"Найдено позиций: {len(results)}")
            back_markup = InlineKeyboardMarkup(
                [[InlineKeyboardButton("← Назад", callback_data="back|root")]]
            )

            lines = []
            for cat, sub, item in results:
                desc = html.escape(item["desc"])
                price = str(item.get("price", "")).strip()
                line = f"<b>{desc}</b>"
                if price:
                    line += f" — <i>{html.escape(price)} ₽</i>"
                line += f"\n<i>{cat} / {sub}</i>"
                lines.extend([line, ""])

            MAX_LEN = 4000
            chunks = []
            cur = ""
            for l in lines:
                seg = l + "\n"
                if len(cur) + len(seg) > MAX_LEN and cur:
                    chunks.append(cur)
                    cur = seg
                else:
                    cur += seg
            if cur:
                chunks.append(cur)

            for idx, chunk in enumerate(chunks):
                if idx == len(chunks) - 1:
                    await update.message.reply_text(chunk, parse_mode="HTML", reply_markup=back_markup)
                else:
                    await update.message.reply_text(chunk, parse_mode="HTML")
            return

        full_catalog = get_full_catalog(context)
        if not full_catalog:
            await update.message.reply_text("Каталог пока не загружен. Пожалуйста, попробуйте позже.")
            return

        # 3) Собираем результаты
        results: list[tuple[str, str, dict]] = []
        brand_subs = {sub.lower() for subs in full_catalog.values() for sub in subs}
        if q in brand_subs:
            for cat, subs in full_catalog.items():
                for sub, items in subs.items():
                    if sub.lower() == q:
                        for item in items:
                            results.append((cat, sub, item))
        else:
            matched_cats = [
                cat for cat in full_catalog
                if cat.lower() == q or cat.lower().startswith(q) or q.startswith(cat.lower())
            ]
            if matched_cats:
                for cat in matched_cats:
                    for sub, items in full_catalog[cat].items():
                        for item in items:
                            results.append((cat, sub, item))
            else:
                for cat, subs in full_catalog.items():
                    for sub, items in subs.items():
                        for item in items:
                            desc = str(item.get("desc", "")).lower()
                            d = re.sub(r'([a-zа-яё])(\d)', r'\1 \2', desc)
                            d = re.sub(r'(\d)([a-zа-яё])', r'\1 \2', d)
                            if q in d:
                                results.append((cat, sub, item))

        if not results:
            await update.message.reply_text("Ничего не найдено по вашему запросу.")
            return

        await update.message.reply_text(f"Найдено позиций: {len(results)}")
        back_markup = InlineKeyboardMarkup(
            [[InlineKeyboardButton("← Назад", callback_data="back|root")]]
        )

        lines = []
        for cat, sub, item in results:
            desc = html.escape(str(item["desc"]))
            price = str(item.get("price", "")).strip()
            line = f"<b>{desc}</b>"
            if price:
                line += f" — <i>{html.escape(price)} ₽</i>"
            line += f"\n<i>{cat} / {sub}</i>"
            lines.extend([line, ""])

        MAX_LEN = 4000
        chunks = []
        current = ""
        for l in lines:
            segment = l + "\n"
            if len(current) + len(segment) > MAX_LEN and current:
                chunks.append(current)
                current = segment
            else:
                current += segment
        if current:
            chunks.append(current)

        for idx, chunk in enumerate(chunks):
            if idx == len(chunks) - 1:
                await update.message.reply_text(chunk, parse_mode="HTML", reply_markup=back_markup)
            else:
                await update.message.reply_text(chunk, parse_mode="HTML")
        return

    # --- Шаг 3.1: парсим номера строк для переноса ---
    if context.user_data.get("change_step") == "awaiting_selection":
        raw = (update.message.text or "").strip()
        import re
        parts = re.split(r"[,\s]+", raw)
        idxs = set()
        try:
            for part in parts:
                if not part:
                    continue
                if "-" in part:
                    a, b = map(int, part.split("-", 1))
                    if a > b:
                        a, b = b, a
                    idxs.update(range(a, b + 1))
                else:
                    idxs.add(int(part))
        except Exception:
            await update.message.reply_text("Некорректный формат. Пример: 1-3,6")
            return

        zero_based = sorted({i - 1 for i in idxs if i > 0})
        sel_map = context.user_data.get("change_selection_map") or []
        picks = []
        for i in zero_based:
            if 0 <= i < len(sel_map):
                picks.append(sel_map[i])

        if not picks:
            await update.message.reply_text("Ничего не выбрано. Укажите корректные номера.")
            return

        # сохраняем конкретные выбранные элементы (источник + индекс + desc/price для надёжного совпадения)
        context.user_data["change_picks"] = picks
    
        # Переходим к выбору новой категории (из полного каталога)
        context.user_data["change_step"] = "awaiting_new_cat"
        full = get_full_catalog(context)
        buttons = [[InlineKeyboardButton(cat, callback_data=f"newcat|{cat}")] for cat in full.keys()]
        await update.message.reply_text("Выберите <b>новую</b> категорию:", reply_markup=InlineKeyboardMarkup(buttons), parse_mode=ParseMode.HTML)
        return

    # --- 2. Обработка нажатий на основные кнопки ---
    if text == BTN_ADMIN_PANEL and is_admin_user:
        admin_buttons = [
            [InlineKeyboardButton("📥 Добавить каталог (.xlsx)", callback_data="adminpanel_add_catalog")],
            [InlineKeyboardButton("🔀 Изменить категорию товаров", callback_data="adminpanel_change_category")],
            [InlineKeyboardButton("📝 Ручные (manual)", callback_data="adminpanel_manual_root")],
            [InlineKeyboardButton("👤 Управление администраторами", callback_data="adminpanel_edit_admins")],
        ]
        markup = InlineKeyboardMarkup(admin_buttons)
        await update.message.reply_text("Админ-панель:", reply_markup=markup)
        return

    if text == BTN_SEARCH_CATALOG:
        context.user_data["awaiting_search"] = True
        await update.message.reply_text("Введите поисковый запрос по каталогу:")
        return

    if text == BTN_CHOOSE_CATEGORY:
        full_catalog = get_full_catalog(context)
        if full_catalog:
            buttons = []
            for cat_name in _sort_categories(list(full_catalog.keys())):
                subdict = full_catalog[cat_name]
                count = sum(len(items) for items in subdict.values())
                buttons.append([InlineKeyboardButton(
                    text=f"{cat_name} ({count})",
                    callback_data=f"cat|{cat_name}"
                )])
            markup = InlineKeyboardMarkup(buttons)
            await update.message.reply_text("Выберите категорию:", reply_markup=markup)
        else:
            await update.message.reply_text("Каталог пока не загружен. Пожалуйста, попробуйте позже.")
        return

    if text == BTN_CONTACT_MANAGER:
        link_btn_tg = InlineKeyboardButton("Написать менеджеру в Телеграм", url=MANAGER_TELEGRAM_LINK)
        link_btn_wa = InlineKeyboardButton("Написать менеджеру в WhatsApp", url=MANAGER_WHATSAPP_LINK)
        await update.message.reply_text(
            "Выберите удобный способ связи с нашим менеджером:",
            reply_markup=InlineKeyboardMarkup([[link_btn_tg], [link_btn_wa]]),
        )
        return

    elif text == BTN_GET_EXCEL:
        import pandas as pd
        import tempfile, os
    
        full_catalog = get_full_catalog(context)
    
        # 1) Собираем строки под требуемые столбцы xmlid/description/price
        rows = []
        for cat, subdict in full_catalog.items():
            for sub, items in subdict.items():
                for item in items:
                    rows.append({
                        "xmlid": f"{cat}/{sub}",                          # Категория/Подкатегория
                        "description": str(item.get("desc", "")),         # Описание
                        "price": item.get("price", "")                    # Цена (преобразуем ниже в число)
                    })
    
        if not rows:
            await update.message.reply_text("Каталог пуст.")
            return
    
        # 2) DataFrame в нужном порядке столбцов
        df = pd.DataFrame(rows, columns=["xmlid", "description", "price"])
    
        # 3) Приводим price к числовому виду (int), вытаскивая только цифры
        def _to_int(v):
            s = str(v)
            digits = "".join(ch for ch in s if ch.isdigit())
            return int(digits) if digits else None
    
        df["price"] = df["price"].apply(_to_int)
    
        # 4) Пишем XLSX: сначала пробуем xlsxwriter (лучший контроль форматов), иначе openpyxl
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()
    
        try:
            try:
                import xlsxwriter  # если установлен — используем
    
                with pd.ExcelWriter(tmp_path, engine="xlsxwriter") as writer:
                    sheet_name = "catalog"
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
    
                    workbook  = writer.book
                    worksheet = writer.sheets[sheet_name]
    
                    # Колонки: 0=xmlid, 1=description, 2=price
                    # Числовой формат для price: #,##0 (будет выглядеть как 75,000)
                    price_fmt = workbook.add_format({"num_format": "#,##0"})
                    worksheet.set_column(0, 0, 24)           # xmlid
                    worksheet.set_column(1, 1, 48)           # description
                    worksheet.set_column(2, 2, 12, price_fmt)  # price (с форматом)
    
            except ImportError:
                # Фолбэк: openpyxl — тоже задаём формат #,##0 для столбца price
                from openpyxl.styles import numbers
                from openpyxl.utils import get_column_letter
    
                with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                    sheet_name = "catalog"
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    ws = writer.sheets[sheet_name]
    
                    # Ширины столбцов
                    ws.column_dimensions[get_column_letter(1)].width = 24   # xmlid
                    ws.column_dimensions[get_column_letter(2)].width = 48   # description
                    ws.column_dimensions[get_column_letter(3)].width = 12   # price
    
                    # Формат для price (колонка C, индекс 3 в 1-based)
                    price_col = 3
                    for row in range(2, len(df) + 2):  # начиная со 2-й строки (после заголовков)
                        cell = ws.cell(row=row, column=price_col)
                        # Только если там число (None/пустые пропускаем)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0"
    
            await update.message.reply_document(document=open(tmp_path, "rb"), filename="catalog.xlsx")
            return
    
        except Exception as exc:
            await update.message.reply_text(f"Не удалось отправить файл: {exc}")
            return
    
        finally:
            try:
                os.remove(tmp_path)
            except Exception:
                pass


    if text == BTN_SUBSCRIBE:
        subs: set[int] = context.application.bot_data.setdefault("subscribers", set())
        if user_id:
            subs.add(user_id)
            await update.message.reply_text("Спасибо! Вы подписаны на обновления.")
        else:
            await update.message.reply_text("Не удалось выполнить подписку.")
        return

    # --- 3. Обработка неизвестных сообщений ---
    await update.message.reply_text(
        "Извините, я вас не понял. Пожалуйста, выберите действие из меню ниже.",
        reply_markup=get_main_menu_markup(is_admin_user),
    )

async def callback_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    data = query.data or ""

    if data == "adminpanel_manual_root":
        submenu = [
            [InlineKeyboardButton("🗂️ Управление категориями", callback_data="adminpanel_edit_category")],
            [InlineKeyboardButton("📦 Управление товарами", callback_data="adminpanel_edit_products")],
            [InlineKeyboardButton("💲 Изменить цены", callback_data="adminpanel_edit_prices")],
            [InlineKeyboardButton("← Назад", callback_data="adminpanel_back")],
        ]
        await query.edit_message_text("Раздел «Ручные (manual_categories.json)»:", reply_markup=InlineKeyboardMarkup(submenu))
        return
    # --- Обработка кнопок админ-панели ---
    if data == "adminpanel_back":
        # Вернуться в главное меню
        user_id = update.effective_user.id if update.effective_user else None
        is_admin_user = user_id and is_admin(user_id)
        await query.edit_message_text("Главное меню:")
        await context.bot.send_message(chat_id=update.effective_chat.id, text="Выберите действие:", reply_markup=get_main_menu_markup(is_admin_user))
        return
    if data == "adminpanel_add_catalog":
        # Выполнить команду /add_catalog
        await add_catalog_command(update, context)
        await query.answer()
        return
    if data == "adminpanel_edit_category":
        await edit_category_command(update, context)
        await query.answer()
        return
    
    if data == "adminpanel_edit_products":
        # 1) Проверяем, что это админ
        user_id = query.from_user.id
        if not is_admin(user_id):
            # шлём алерт, чтобы не мешать основному чату
            await query.answer("Извините, команда доступна только администратору.", show_alert=True)
            return

        # 2) Загружаем вручную добавленные категории/бренды
        manual_cats = context.application.bot_data.get("manual_categories")
        if manual_cats is None:
            manual_cats = _load_manual_categories()
            context.application.bot_data["manual_categories"] = manual_cats

        # 3) Собираем кнопки «Категория / Бренд»
        buttons = []
        cb_map = {}
        idx = 0
        for cat, brands in manual_cats.items():
            for brand in brands:
                key = f"manualprod_select|{idx}"
                cb_map[key] = (cat, brand)
                buttons.append([InlineKeyboardButton(f"{cat} / {brand}", callback_data=key)])
                idx += 1

        if not buttons:
            # Если ничего нет — просто редактируем текст
            await query.edit_message_text("Нет вручную добавленных подкатегорий для управления товарами.")
        else:
            # Сохраняем mapping и показываем клавиатуру
            context.user_data["manualprod_select_map"] = cb_map
            markup = InlineKeyboardMarkup(buttons)
            await query.edit_message_text(
                "Выберите подкатегорию для редактирования товаров:",
                reply_markup=markup
            )

        await query.answer()
        return

    if data == "adminpanel_edit_admins":
        await edit_admins_command(update, context)
        await query.answer()
        return
    
    # --- Новая функция: Изменить категорию товаров ---
    if data == "adminpanel_change_category":
        user_id = query.from_user.id
        if not is_admin(user_id):
            await query.answer("Извините, команда доступна только администратору.", show_alert=True)
            return

        full = get_full_catalog(context)
        if not full:
            await query.edit_message_text("Каталог пуст.")
            return

        # Кнопки категорий с общим количеством позиций (auto + moved + manual)
        buttons = []
        for cat_name in _sort_categories(list(full.keys())):
            subdict = full.get(cat_name, {})
            count = sum(len(items) for items in subdict.values())
            buttons.append([InlineKeyboardButton(f"{cat_name} ({count})", callback_data=f"change|cat|{cat_name}")])

        context.user_data["change_step"] = "awaiting_cat"
        await query.edit_message_text("Выберите категорию, из которой переносим:", reply_markup=InlineKeyboardMarkup(buttons))
        return

    # Шаг 2: после нажатия change|cat|<категория> — выбор подкатегории
    if data.startswith("change|cat|"):
        _, _, cat = data.split("|", 2)

        # Сохраняем исходную категорию
        context.user_data["change_cat"] = cat
        context.user_data["change_step"] = "awaiting_sub"

        auto   = (context.application.bot_data.get("catalog") or {}).get(cat, {}) or {}
        moved  = (context.application.bot_data.get("moved_overrides") or {}).get(cat, {}) or {}
        manual = (context.application.bot_data.get("manual_categories") or {}).get(cat, {}) or {}

        # Объединяем подкатегории и считаем общее количество
        all_subs = sorted(set(auto.keys()) | set(moved.keys()) | set(manual.keys()))
        if not all_subs:
            await query.edit_message_text("В этой категории пока нет подкатегорий.")
            return

        buttons = []
        for sub in all_subs:
            cnt = len(auto.get(sub, [])) + len(moved.get(sub, [])) + len(manual.get(sub, []))
            buttons.append([InlineKeyboardButton(f"{sub} ({cnt})", callback_data=f"change|sub|{cat}|{sub}")])

        await query.edit_message_text(f"Категория: {cat}\nВыберите подкатегорию:", reply_markup=InlineKeyboardMarkup(buttons))
        return


    # Шаг 3: после on change|sub|<cat>|<sub> — показываем список товаров, ждём ввода номеров
    if data.startswith("change|sub|"):
        _, _, cat, sub = data.split("|", 3)

        context.user_data["change_cat"] = cat
        context.user_data["change_sub"] = sub
        context.user_data["change_step"] = "awaiting_selection"

        auto_list   = (context.application.bot_data.get("catalog") or {}).get(cat, {}).get(sub, []) or []
        moved_list  = (context.application.bot_data.get("moved_overrides") or {}).get(cat, {}).get(sub, []) or []
        manual_list = (context.application.bot_data.get("manual_categories") or {}).get(cat, {}).get(sub, []) or []

        # Строим объединённый вывод и map "номер -> источник"
        lines = [f"<b>{cat} / {sub}</b>", "Выберите строки для переноса (например: 1-3,6)", ""]
        selection_map = []  # список словарей: {"src": "auto|moved|manual", "idx": int, "desc": str, "price": str}

        idx = 1
        def _add_block(title, src, lst):
            nonlocal idx, lines, selection_map
            if lst:
                lines.append(f"<i>{title}</i>")
                for i, it in enumerate(lst):
                    d = html.escape(str(it.get("desc", "")))
                    p = html.escape(str(it.get("price", "")))
                    lines.append(f"{idx}. {d} — {p}")
                    selection_map.append({"src": src, "idx": i, "desc": str(it.get("desc","")), "price": str(it.get("price",""))})
                    idx += 1
                lines.append("")

        _add_block("Авто-каталог", "auto", auto_list)
        _add_block("Перенесённые", "moved", moved_list)
        _add_block("Ручные", "manual", manual_list)

        if not selection_map:
            await query.edit_message_text("В этой подкатегории нет товаров.")
            return

        context.user_data["change_selection_map"] = selection_map
        await query.edit_message_text("\n".join(lines), parse_mode=ParseMode.HTML)
        return
    
        # --- Шаг 4: выбор новой категории ---
    if data.startswith("newcat|") and context.user_data.get("change_step") == "awaiting_new_cat":
        _, new_cat = data.split("|",1)
        context.user_data["new_cat"] = new_cat
        full = get_full_catalog(context)
        subs = full.get(new_cat, {})
        buttons = [
            [InlineKeyboardButton(f"{sub} ({len(items)})", callback_data=f"newsub|{new_cat}|{sub}")]
            for sub, items in subs.items()
        ]
        await query.edit_message_text(
            f"*Новая категория:* {new_cat}\nВыберите подкатегорию:",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode=ParseMode.MARKDOWN
        )
        return

    # --- Шаг 5: выбор новой подкатегории и перенос ---
    if data.startswith("newsub|") and context.user_data.get("change_step") == "awaiting_new_cat":
        _, new_cat, new_sub = data.split("|", 2)
    
        # Откуда переносим
        src_cat = context.user_data.pop("change_cat")
        src_sub = context.user_data.pop("change_sub")
        picks   = context.user_data.pop("change_picks", [])
        context.user_data.pop("change_selection_map", None)
        context.user_data.pop("change_step", None)
    
        auto_cat = context.application.bot_data.get("catalog") or {}
        overrides = context.application.bot_data.get("moved_overrides") or _load_moved_overrides()
        manual    = context.application.bot_data.get("manual_categories") or _load_manual_categories()
    
        moved_cnt = 0
    
        # Утилита: безопасное удаление конкретного элемента по desc+price
        def _remove_by_desc_price(lst, desc, price):
            for j, it in enumerate(lst):
                if str(it.get("desc","")) == desc and str(it.get("price","")) == price:
                    lst.pop(j)
                    return True
            return False
    
        # 1) Обрабатываем авто-товары: auto -> moved_overrides (с orig_cat/sub)
        auto_list = auto_cat.get(src_cat, {}).get(src_sub, [])
        for pick in [p for p in picks if p["src"] == "auto"]:
            desc, price = pick["desc"], pick["price"]
            if _remove_by_desc_price(auto_list, desc, price):
                overrides.setdefault(new_cat, {}).setdefault(new_sub, []).append({
                    "desc": desc,
                    "price": price,
                    "origin": "auto",
                    "orig_cat": src_cat,
                    "orig_sub": src_sub,
                })
                moved_cnt += 1
    
        # 2) Обрабатываем перенесённые: moved_overrides -> moved_overrides (orig_* не меняем)
        moved_list = overrides.get(src_cat, {}).get(src_sub, [])
        moved_to_keep = []
        for it in moved_list:
            # выясняем, выбран ли этот элемент
            chosen = any(p["src"] == "moved" and p["desc"] == str(it.get("desc","")) and p["price"] == str(it.get("price","")) for p in picks)
            if chosen:
                overrides.setdefault(new_cat, {}).setdefault(new_sub, []).append(it)  # переносим как есть
                moved_cnt += 1
            else:
                moved_to_keep.append(it)
        if moved_list is not None:
            # обновляем/удаляем исходную ветку только если она существует
            if src_cat in overrides and src_sub in overrides[src_cat]:
                if moved_to_keep:
                    overrides[src_cat][src_sub] = moved_to_keep
                else:
                    del overrides[src_cat][src_sub]
                    if not overrides[src_cat]:
                        del overrides[src_cat]
    
        # 3) Обрабатываем ручные: manual -> manual
        manual_list = manual.get(src_cat, {}).get(src_sub, [])
        manual_to_keep = []
        for it in manual_list:
            chosen = any(p["src"] == "manual" and p["desc"] == str(it.get("desc","")) and p["price"] == str(it.get("price","")) for p in picks)
            if chosen:
                manual.setdefault(new_cat, {}).setdefault(new_sub, []).append(it)
                moved_cnt += 1
            else:
                manual_to_keep.append(it)
        if manual_list is not None:
            if manual_to_keep:
                manual[src_cat][src_sub] = manual_to_keep
            else:
                if src_cat in manual and src_sub in manual[src_cat]:
                    del manual[src_cat][src_sub]
                    if not manual[src_cat]:
                        del manual[src_cat]
    
        # Сохраняем изменения
        context.application.bot_data["catalog"] = auto_cat
        _save_catalog_to_disk(auto_cat)
        context.application.bot_data["moved_overrides"] = overrides
        _save_moved_overrides(overrides)
        context.application.bot_data["manual_categories"] = manual
        _save_manual_categories(manual)
    
        await query.edit_message_text(
            f"✅ Перенесено позиций: {moved_cnt}\n"
            f"Из: {src_cat}/{src_sub} → В: {new_cat}/{new_sub}"
        )
        # Возврат в новую админ-панель
        await show_admin_panel(query, context)
        return

    # --- Изменить цены (ручные товары) ---
    if data == "adminpanel_edit_prices":
        user_id = query.from_user.id
        if not is_admin(user_id):
            await query.answer("Доступ только для администратора.", show_alert=True)
            return

        manual = context.application.bot_data.get("manual_categories")
        if manual is None:
            manual = _load_manual_categories()
            context.application.bot_data["manual_categories"] = manual

        buttons = []
        cb_map = {}
        idx = 0
        for cat, brands in manual.items():
            for brand in brands.keys():
                key = f"manualprice_select|{idx}"
                cb_map[key] = (cat, brand)
                buttons.append([InlineKeyboardButton(f"{cat} / {brand}", callback_data=key)])
                idx += 1

        if not buttons:
            await query.edit_message_text("Нет вручную добавленных подкатегорий (manual_categories.json).")
            return

        context.user_data["manualprice_select_map"] = cb_map
        await query.edit_message_text(
            "Выберите подкатегорию для изменения цен:",
            reply_markup=InlineKeyboardMarkup(buttons)
        )
        return
    
    if data.startswith("manualprice_select|"):
        cb_map = context.user_data.get("manualprice_select_map", {})
        if data not in cb_map:
            await query.edit_message_text("Подкатегория не найдена.")
            return

        cat, brand = cb_map[data]
        context.user_data["manualprice_cat"] = cat
        context.user_data["manualprice_brand"] = brand
        context.user_data["manualprice_step"] = "awaiting_indices"

        manual = context.application.bot_data.get("manual_categories", {}) or _load_manual_categories()
        items = manual.get(cat, {}).get(brand, [])

        if not items:
            await query.edit_message_text(f"В {cat} / {brand} товаров нет.")
            return

        # Нумерованный список
        lines = ["<b>Текущие товары:</b>"]
        for i, it in enumerate(items, start=1):
            d = html.escape(it.get("desc", ""))
            p = html.escape(str(it.get("price", "")))
            lines.append(f"{i}. {d} — {p}")
        lines.append("")
        lines.append("Введите номера строк для изменения цены (например: 1-3,5):")

        await query.edit_message_text("\n".join(lines), parse_mode=ParseMode.HTML)
        return

    
    # --- Обработка кнопок для управления вручную добавленными категориями ---
    if data == "manualcat_add":
        context.user_data["manualcat_step"] = 1
        await query.edit_message_text(
        "Введите название категории:\n\n"
        "Для отмены введите /start"
    )
        return
    
    if data == "manualcat_remove":
        # Показываем список для удаления, используем mapping для точного соответствия
        import urllib.parse
        manual_cats = context.application.bot_data.get("manual_categories")
        if manual_cats is None:
            manual_cats = _load_manual_categories()
            context.application.bot_data["manual_categories"] = manual_cats
        buttons = []
        cb_map = {}  # callback_data -> (cat, brand)
        idx = 0
        for cat, brands in manual_cats.items():
            for brand in brands:
                cb_data = f"manualcat_del|{idx}"
                cb_map[cb_data] = (cat, brand)
                btn_text = f"{cat} / {brand}"
                buttons.append([InlineKeyboardButton(btn_text, callback_data=cb_data)])
                idx += 1
        if not buttons:
            await query.edit_message_text("Нет вручную добавленных категорий для удаления.")
            return
        # Сохраняем mapping в user_data
        context.user_data["manualcat_del_map"] = cb_map
        markup = InlineKeyboardMarkup(buttons)
        await query.edit_message_text("Выберите категорию/бренд для удаления:", reply_markup=markup)
        return
    
    if data.startswith("manualcat_del|"):
        cb_map = context.user_data.get("manualcat_del_map", {})
        if data not in cb_map:
            await query.edit_message_text("Категория/бренд не найдены.")
            return

        cat, brand = cb_map[data]

        # 1) Удаляем ручную подкатегорию из manual_categories
        manual_cats = context.application.bot_data.get("manual_categories")
        if manual_cats is None:
            manual_cats = _load_manual_categories()

        removed_manual_count = 0
        if cat in manual_cats and brand in manual_cats[cat]:
            removed_manual_count = len(manual_cats[cat][brand])
            del manual_cats[cat][brand]
            if not manual_cats[cat]:
                del manual_cats[cat]
            context.application.bot_data["manual_categories"] = manual_cats
            _save_manual_categories(manual_cats)

        # 2) Если в этой же подкатегории лежали ПЕРЕНЕСЁННЫЕ товары (moved_overrides) — вернём их в исходные места
        overrides = context.application.bot_data.get("moved_overrides")
        if overrides is None:
            overrides = _load_moved_overrides()

        returned_count = 0
        if overrides.get(cat, {}).get(brand):
            moved_items = overrides[cat][brand]
            catalog = context.application.bot_data.get("catalog") or {}

            for it in moved_items:
                desc = it.get("desc", "")
                price = it.get("price", "")
                o_cat = it.get("orig_cat")
                o_sub = it.get("orig_sub")
                if not o_cat or not o_sub:
                    # на случай старых записей без orig_* — пробуем классифицировать по описанию
                    o_cat, o_sub = extract_category(desc)

                catalog.setdefault(o_cat, {}).setdefault(o_sub, []).append({"desc": desc, "price": price})
                returned_count += 1

            # Удаляем перенесённые из этой ручной подкатегории
            del overrides[cat][brand]
            if not overrides[cat]:
                del overrides[cat]

            # Сохраняем обе структуры
            context.application.bot_data["catalog"] = catalog
            _save_catalog_to_disk(catalog)
            context.application.bot_data["moved_overrides"] = overrides
            _save_moved_overrides(overrides)

        # 3) Ответ и возврат в актуальную админ-панель
        await query.edit_message_text(
            f"Удалено: {cat} / {brand}\n"
            f"Возвращено в исходные категории: {returned_count} поз."
        )
        context.user_data.pop("manualcat_del_map", None)
        await show_admin_panel(query, context)
        return
    
    # --- Обработка кнопок для управления админами ---
    if data == "admin_add":
        context.user_data["awaiting_admin_action"] = "add"
        await query.edit_message_text("Введите user_id пользователя, которого нужно добавить в администраторы:")
        return
    if data == "admin_remove":
        # Показываем список админов с кнопками для удаления
        admins = _load_admins()
        buttons = []
        for admin_id in admins:
            try:
                user = await context.bot.get_chat(admin_id)
                username = f"@{user.username}" if getattr(user, "username", None) else ""
            except Exception:
                username = ""
            btn_text = f"{admin_id} {username}".strip()
            buttons.append([InlineKeyboardButton(btn_text, callback_data=f"admin_del|{admin_id}")])
        if not buttons:
            await query.edit_message_text("Нет администраторов для удаления.")
            return
        markup = InlineKeyboardMarkup(buttons)
        await query.edit_message_text("Выберите администратора для удаления:", reply_markup=markup)
        return
    if data.startswith("admin_del|"):
        # Удаляем выбранного админа
        parts = data.split("|", 1)
        if len(parts) == 2:
            try:
                target_id = int(parts[1])
            except Exception:
                await query.edit_message_text("Некорректный user_id.")
                return
            admins = _load_admins()
            if target_id in admins:
                admins.remove(target_id)
                _save_admins(admins)
                await query.edit_message_text(f"Пользователь {target_id} удалён из администраторов.")
                await show_admin_panel(update, context)
            else:
                await query.edit_message_text("Такого пользователя нет в списке админов.")
        return
    await query.answer()
    parts = data.split("|")
    if not parts:
        return
    
        # --- Обработка выбора вручную добавленной подкатегории для редактирования товаров ---
    if data.startswith("manualprod_select|"):
        cb_map = context.user_data.get("manualprod_select_map", {})
        if data not in cb_map:
            await query.edit_message_text("Подкатегория не найдена.")
            return
        cat, brand = cb_map[data]
        context.user_data["manualprod_cat"] = cat
        context.user_data["manualprod_brand"] = brand

        # Берём существующие товары
        manual_cats = context.application.bot_data.get("manual_categories", {})
        items = manual_cats.get(cat, {}).get(brand, [])

        # Формируем список в текстовом виде
        if items:
            lines = ["<b>Текущие товары:</b>"]
            for idx, it in enumerate(items, start=1):
                desc = html.escape(it.get("desc", ""))
                price = html.escape(str(it.get("price", "")))
                lines.append(f"{idx}. {desc} — {price}")
            lines.append("")  # пустая строка перед кнопками
        else:
            lines = ["<i>Товаров ещё нет.</i>", ""]

        # Кнопки действий
        buttons = [
            [InlineKeyboardButton("Добавить товары", callback_data="manualprod_add")],
            [InlineKeyboardButton("Удалить товары", callback_data="manualprod_remove")],
        ]

        await query.edit_message_text(
            "\n".join(lines) +
            f"\nПодкатегория <b>{cat} / {brand}</b>\nЧто вы хотите сделать?",
            reply_markup=InlineKeyboardMarkup(buttons),
            parse_mode="HTML"
        )
        return

    # --- Начало добавления товаров ---
    if data == "manualprod_add":
        # Устанавливаем шаг: ожидание списка товаров
        context.user_data["manualprod_step"] = 1
        await query.edit_message_text(
            "Введите описание товара и цену.\nКаждая строка: Описание;Цена\n\nДля отмены введите /start"
        )
        return

       # --- Начало удаления товаров: ввод номеров строк ---
    if data == "manualprod_remove":
        cat = context.user_data.get("manualprod_cat")
        brand = context.user_data.get("manualprod_brand")
        manual_cats = context.application.bot_data.get("manual_categories", {})
        items = manual_cats.get(cat, {}).get(brand, [])
        if not items:
            await query.edit_message_text("Товаров для удаления нет.")
            return

        # Формируем нумерованный список
        lines = []
        for idx, it in enumerate(items, start=1):
            desc = html.escape(it.get("desc", ""))
            price = html.escape(str(it.get("price", "")))
            lines.append(f"{idx}. {desc} — {price}")
        text = "<b>Товары для удаления:</b>\n\n" + "\n".join(lines)
        await query.edit_message_text(
            text + "\n\nНапишите номера строк для удаления (например: 1-3,5):",
            parse_mode=ParseMode.HTML
        )

        # Переходим к шагу парсинга
        context.user_data["manualprod_step"] = "awaiting_manualprod_delete"
        return

    # --- Удаление выбранного товара ---
    if data.startswith("manualprod_del|"):
        cb_map = context.user_data.get("manualprod_del_map", {})
        if data not in cb_map:
            await query.edit_message_text("Товар не найден.")
            return
        idx = cb_map[data]
        cat = context.user_data.get("manualprod_cat")
        brand = context.user_data.get("manualprod_brand")
        manual_cats = context.application.bot_data.get("manual_categories", {})
        items = manual_cats.get(cat, {}).get(brand, [])
        if 0 <= idx < len(items):
            deleted = items.pop(idx)
            # Сохраняем изменения
            _save_manual_categories(manual_cats)
            context.application.bot_data["manual_categories"] = manual_cats
            await query.edit_message_text(f"Удалён товар: {deleted.get('desc')} — {deleted.get('price')}")
            await show_admin_panel(update, context)
        else:
            await query.edit_message_text("Некорректный индекс.")
        return


    full_catalog = get_full_catalog(context)
    if not full_catalog:
        await query.edit_message_text("Каталог не найден. Загрузите файл командой /add_catalog.")
        return

    if parts[0] == "cat":  # Выбрана категория
        cat = parts[1]
        # Навигационный стек: пушим текущий уровень
        nav_stack = context.user_data.get("navigation_stack", [])
        # Если пришли не из back, пушим
        if not nav_stack or nav_stack[-1] != ("cat", cat):
            nav_stack.append(("cat", cat))
        context.user_data["navigation_stack"] = nav_stack
        subcats = full_catalog.get(cat, {})
        # Кнопки подкатегорий с количеством товаров
        buttons = []
        for sub_name, items in subcats.items():
            buttons.append([InlineKeyboardButton(text=f"{sub_name} ({len(items)})", callback_data=f"sub|{cat}|{sub_name}")])
        # Кнопка назад: если стек не пуст, возвращаемся к предыдущему уровню
        if len(nav_stack) > 1:
            buttons.append([InlineKeyboardButton(text="← Назад", callback_data="back")])
        else:
            buttons.append([InlineKeyboardButton(text="← Назад", callback_data="back|root")])
        markup = InlineKeyboardMarkup(buttons)
        await query.edit_message_text(f"Категория: {cat}\nВыберите подкатегорию:", reply_markup=markup)
        return

    elif parts[0] == "sub":  # Выбрана подкатегория
        cat, sub = parts[1], parts[2]
        # Навигационный стек: пушим текущий уровень
        nav_stack = context.user_data.get("navigation_stack", [])
        if not nav_stack or nav_stack[-1] != ("sub", cat, sub):
            nav_stack.append(("sub", cat, sub))
        context.user_data["navigation_stack"] = nav_stack
        items = full_catalog.get(cat, {}).get(sub, [])

        text_lines: list[str] = []
        for item in items:
            desc = html.escape(str(item['desc']))
            price = str(item['price']).strip()
            line = f"<b>{desc}</b>"
            if price:
                line += f" — <i>{html.escape(price)} ₽</i>"
            text_lines.append(line)
        # Добавляем пустую строку между товарами для читаемости
        lines_with_spacing = []
        for l in text_lines:
            lines_with_spacing.append(l)
            lines_with_spacing.append("")

        MAX_LENGTH = 4000
        chunks: list[str] = []
        current_lines: list[str] = []
        current_len = 0
        for line in lines_with_spacing:
            line_len = len(line) + 1
            if current_len + line_len > MAX_LENGTH and current_lines:
                chunks.append("\n".join(current_lines))
                current_lines = [line]
                current_len = line_len
            else:
                current_lines.append(line)
                current_len += line_len
        if current_lines:
            chunks.append("\n".join(current_lines))

        if not chunks:
            chunks = ["Нет товаров."]

        # Кнопка назад: если стек не пуст, возвращаемся к предыдущему уровню
        nav_stack = context.user_data.get("navigation_stack", [])
        if len(nav_stack) > 1:
            buttons = [[InlineKeyboardButton(text="← Назад", callback_data="back")]]
        else:
            buttons = [[InlineKeyboardButton(text="← Назад", callback_data="back|root")]]
        markup = InlineKeyboardMarkup(buttons)

        # Отправляем все чанки как отдельные сообщения, как при поиске
        # Если только один чанк — показываем кнопку "Назад" сразу
        n = len(chunks)
        if n == 1:
            text_to_send = f"Категория: {cat} / {sub}\n\n{chunks[0]}"
            await query.edit_message_text(text_to_send, parse_mode="HTML", reply_markup=markup)
        else:
            for idx, chunk in enumerate(chunks):
                if idx == 0:
                    # Первый чанк — только заголовок
                    text_to_send = f"Категория: {cat} / {sub}\n\n{chunk}"
                    await query.edit_message_text(text_to_send, parse_mode="HTML")
                elif idx < n - 1:
                    await context.bot.send_message(chat_id=update.effective_chat.id, text=chunk, parse_mode="HTML")
                else:
                    # Последний чанк — с кнопкой "Назад"
                    await context.bot.send_message(chat_id=update.effective_chat.id, text=chunk, parse_mode="HTML", reply_markup=markup)
        return

    elif parts[0] == "back":
        # Навигационный стек: pop текущий уровень
        nav_stack = context.user_data.get("navigation_stack", [])
        if nav_stack:
            nav_stack.pop()
        context.user_data["navigation_stack"] = nav_stack

        # Если стек пуст или явно back|root — показываем корень каталога

        if (len(parts) > 1 and parts[1] == "root") or not nav_stack:
            buttons = []
            for cat_name in _sort_categories(list(full_catalog.keys())):
                subdict = full_catalog[cat_name]
                count = sum(len(items) for items in subdict.values())
                buttons.append([InlineKeyboardButton(text=f"{cat_name} ({count})", callback_data=f"cat|{cat_name}")])
            markup = InlineKeyboardMarkup(buttons)
            try:
                await query.edit_message_text("Выберите категорию:", reply_markup=markup)
            except Exception as e:
                from telegram.error import BadRequest
                if isinstance(e, BadRequest):
                    await context.bot.send_message(chat_id=update.effective_chat.id, text="Выберите категорию:", reply_markup=markup)
                else:
                    raise
            return

        # Иначе — показываем предыдущий уровень
        prev = nav_stack[-1] if nav_stack else None
        if prev:
            if prev[0] == "cat":
                cat = prev[1]
                subcats = full_catalog.get(cat, {})
                buttons = []
                for sub_name, items in subcats.items():
                    buttons.append([InlineKeyboardButton(text=f"{sub_name} ({len(items)})", callback_data=f"sub|{cat}|{sub_name}")])
                if len(nav_stack) > 1:
                    buttons.append([InlineKeyboardButton(text="← Назад", callback_data="back")])
                else:
                    buttons.append([InlineKeyboardButton(text="← Назад", callback_data="back|root")])
                markup = InlineKeyboardMarkup(buttons)
                await query.edit_message_text(f"Категория: {cat}\nВыберите подкатегорию:", reply_markup=markup)
            elif prev[0] == "sub":
                cat, sub = prev[1], prev[2]
                items = full_catalog.get(cat, {}).get(sub, [])
                text_lines: list[str] = []
                for item in items:
                    desc = html.escape(str(item['desc']))
                    price = str(item['price']).strip()
                    line = f"<b>{desc}</b>"
                    if price:
                        line += f" — <i>{html.escape(price)} ₽</i>"
                    text_lines.append(line)
                # Добавляем пустую строку между товарами для читаемости
                lines_with_spacing = []
                for l in text_lines:
                    lines_with_spacing.append(l)
                    lines_with_spacing.append("")

                MAX_LENGTH = 4000
                chunks: list[str] = []
                current_lines: list[str] = []
                current_len = 0
                for line in lines_with_spacing:
                    line_len = len(line) + 1
                    if current_len + line_len > MAX_LENGTH and current_lines:
                        chunks.append("\n".join(current_lines))
                        current_lines = [line]
                        current_len = line_len
                    else:
                        current_lines.append(line)
                        current_len += line_len
                if current_lines:
                    chunks.append("\n".join(current_lines))
                if not chunks:
                    chunks = ["Нет товаров."]
                if len(nav_stack) > 1:
                    buttons = [[InlineKeyboardButton(text="← Назад", callback_data="back")]]
                else:
                    buttons = [[InlineKeyboardButton(text="← Назад", callback_data="back|root")]]
                markup = InlineKeyboardMarkup(buttons)
                text_to_send = f"Категория: {cat} / {sub}\n\n{chunks[0]}"
                await query.edit_message_text(text_to_send, reply_markup=markup, parse_mode="HTML")
        return

def main() -> None:
    """Запуск бота."""
    if TOKEN == "YOUR_BOT_TOKEN_HERE":
        raise RuntimeError(
            "Необходимо задать токен Telegram-бота. "
            "Отредактируйте переменную TOKEN или задайте TG_BOT_TOKEN."
        )

    app = ApplicationBuilder().token(TOKEN).build()

    # Загружаем каталог с диска при старте и сохраняем в bot_data
    initial_catalog = _load_catalog_from_disk()
    if initial_catalog:
        app.bot_data["catalog"] = initial_catalog

    # Загружаем вручную добавленные категории с диска
    app.bot_data["manual_categories"] = _load_manual_categories()
    app.bot_data["moved_overrides"] = _load_moved_overrides()

    # Регистрируем обработчики
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add_catalog", add_catalog_command))
    app.add_handler(CommandHandler("edit_category", edit_category_command))
    app.add_handler(  CommandHandler("edit_products", edit_products_command) ) 
    app.add_handler(CommandHandler("edit_admins", edit_admins_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("about", about_command))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & (~filters.COMMAND) & (~filters.Document.ALL), handle_text))
    app.add_handler(CallbackQueryHandler(callback_query_handler))

    # Запускаем бесконечный поллинг
    print("Бот запущен. Нажмите Ctrl-C для остановки.")
    app.run_polling()


if __name__ == "__main__":
    main() 